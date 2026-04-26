# -*- coding: utf-8 -*-

from pathlib import Path
from .base import DocumentParser

class ExcelParser(DocumentParser):
    def is_supported(self, file_path: str) -> bool:
        return Path(file_path).suffix.lower() == ".xlsx"

    def parse(self, file_path: str) -> str:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        content = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            content.append(f"=== Sheet: {sheet_name} ===")
            for row in sheet.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                content.append(" | ".join(row_data))
        return "\n".join(content)